"""Excel and CSV export service."""
import csv
import io
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


HEADER_FILL = "6C63FF"
ALT_FILL    = "F0F2FF"


def _style_header(ws, row: int, cols: int):
    if not OPENPYXL_AVAILABLE:
        return
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    if not OPENPYXL_AVAILABLE:
        return
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_members_excel(members: List[Dict[str, Any]]) -> bytes:
    if not OPENPYXL_AVAILABLE:
        return export_members_csv(members).encode()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    headers = ["Member ID", "First Name", "Last Name", "Email", "Phone", "Status", "Plan", "Check-ins", "Lifetime Value", "Joined", "Expires"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20
    for i, m in enumerate(members, start=2):
        ws.append([
            m.get("member_id"), m.get("first_name"), m.get("last_name"),
            m.get("email"), m.get("phone"), m.get("status"),
            "", str(m.get("total_checkins", 0)),
            str(m.get("lifetime_value", 0)),
            str(m.get("created_at", ""))[:10], "",
        ])
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = PatternFill(start_color=ALT_FILL, end_color=ALT_FILL, fill_type="solid")
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_members_csv(members: List[Dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "member_id","first_name","last_name","email","phone","status","total_checkins","lifetime_value","created_at"
    ])
    writer.writeheader()
    for m in members:
        writer.writerow({k: m.get(k, "") for k in writer.fieldnames})
    return output.getvalue()


def export_invoices_excel(invoices: List[Dict[str, Any]]) -> bytes:
    if not OPENPYXL_AVAILABLE:
        return export_invoices_csv(invoices).encode()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = ["Invoice #", "Member ID", "Description", "Subtotal", "Discount", "Tax", "Total", "Paid", "Due", "Status", "Due Date", "Paid At"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for i, inv in enumerate(invoices, start=2):
        ws.append([
            inv.get("invoice_number"), inv.get("member_id"),
            inv.get("description", ""),
            float(inv.get("subtotal", 0)), float(inv.get("discount_amount", 0)),
            float(inv.get("tax_amount", 0)), float(inv.get("total", 0)),
            float(inv.get("amount_paid", 0)), float(inv.get("amount_due", 0)),
            inv.get("status"), str(inv.get("due_date", ""))[:10],
            str(inv.get("paid_at", "") or "")[:10],
        ])
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_invoices_csv(invoices: List[Dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "invoice_number","member_id","subtotal","discount_amount","tax_amount",
        "total","amount_paid","amount_due","status","due_date","paid_at"
    ])
    writer.writeheader()
    for inv in invoices:
        writer.writerow({k: inv.get(k, "") for k in writer.fieldnames})
    return output.getvalue()


def export_attendance_csv(logs: List[Dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "id","member_id","branch_id","check_in","check_out","method","duration_minutes"
    ])
    writer.writeheader()
    for log in logs:
        cin = log.get("check_in")
        cout = log.get("check_out")
        duration = None
        if cin and cout:
            from datetime import datetime
            try:
                duration = int((datetime.fromisoformat(cout) - datetime.fromisoformat(cin)).total_seconds() / 60)
            except Exception:
                pass
        writer.writerow({
            "id": log.get("id"), "member_id": log.get("member_id"),
            "branch_id": log.get("branch_id"), "check_in": cin, "check_out": cout,
            "method": log.get("method"), "duration_minutes": duration,
        })
    return output.getvalue()


def export_payroll_excel(slips: List[Dict[str, Any]], run: Dict[str, Any]) -> bytes:
    if not OPENPYXL_AVAILABLE:
        return b""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.get('period_start','')}"
    headers = ["Staff ID","Base Salary","Days Worked","Days Absent","OT Hours","OT Pay","Commission","Bonus","Gross","GOSI","Absence Ded.","Other Ded.","Total Ded.","NET PAY","Paid"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    total_net = 0
    for i, s in enumerate(slips, start=2):
        net = float(s.get("net", 0))
        total_net += net
        ws.append([
            s.get("staff_id"), float(s.get("base_salary",0)),
            s.get("days_worked",0), s.get("days_absent",0),
            float(s.get("overtime_hours",0)), float(s.get("overtime_pay",0)),
            float(s.get("commission",0)), float(s.get("bonus",0)),
            float(s.get("gross",0)), float(s.get("deduction_gosi",0)),
            float(s.get("deduction_absence",0)), float(s.get("deduction_other",0)),
            float(s.get("total_deductions",0)), net,
            "Yes" if s.get("is_paid") else "No",
        ])
    # Summary row
    ws.append(["TOTAL","","","","","","","","","","","","", total_net, ""])
    last = ws.max_row
    for col in range(1, len(headers)+1):
        ws.cell(row=last, column=col).font = Font(bold=True)
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
